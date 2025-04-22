import openpyxl

# openpyxl.load_workbook("inventory.xlsx")

# Save it as a variable

inventory_file = openpyxl.load_workbook("inventory.xlsx")

product_list = inventory_file.active

# First task is to calculate how many products I have per supplier and then list the names of the supplier with that respective number of products 

  # Dictionary : Key is Name of company, Value is a number of that company have

products_per_supplier = {}

# Second Task is to calculate total inventory value per supplier meaning in my lists for each product I have number of units for the product and the price so basically for each supplier 
 
total_value_per_supplier = {}

# Third task is to print out all the products that have inventory less than 10 . 

product_inventory_under_10 = {}

  # Loop through those spreadsheet role to get Name and number of product and start from second row
    # range() will create a list of number to iterate through 

# Fourth task is to add some value inside the spreadsheet

for product_row in range(2 ,product_list.max_row + 1):
  # Get supplier name . cell(row, column)
  supplier_name = (product_list.cell(product_row, 4)).value

  # Get a Inventory list from a Spreadsheet 
  inventory = (product_list.cell(product_row, 2)).value 

  # Get price from a Spreadsheet

  price = (product_list.cell(product_row, 3)).value

  # Get product list 
  product = (product_list.cell(product_row, 1)).value 

  # Add the 5th column to spreadsheet 
  # Without .value I will have the whole cell Object
  inventory_price = product_list.cell(product_row, 5)

  # Add supplier name to Dict . Syntax : products_per_supplier["key"] = "value" 
  # Check if Supplier not in the Dict, Add to Dict with value = 1 
  # If Supplier is already in the Dict, Increase value by 1 
  if supplier_name in products_per_supplier:
    products_per_supplier[supplier_name] = products_per_supplier[supplier_name] + 1
  else:
    products_per_supplier[supplier_name] = 1

  # Second task Calculation total value of the inventory per supplier 
  # Get a total_price_of_inventory by product inventory * product price 
  total_price_of_inventory = (product_list.cell(product_row, 2)).value * (product_list.cell(product_row, 3)).value

  # If supplier name in total_value_per_supplier get that supplier name value + total_price_of_inventory 
  # If not add total_price_of_inventory to total_value_per_supplier
  if supplier_name in total_value_per_supplier:
    total_value_per_supplier[supplier_name] = total_price_of_inventory + total_value_per_supplier[supplier_name]
  else:
    total_value_per_supplier[supplier_name] = total_price_of_inventory

  # Logic product with inventory less than 10 

  if inventory < 10:
    product_inventory_under_10[int(product)] = int(inventory)

  # Add value for total Inventory price 
 
  inventory_price.value = inventory * price

print(products_per_supplier)

print(total_value_per_supplier)

print (product_inventory_under_10)
# To explicitly save the file 
# This line will save to change I create a new one from the original one 
inventory_file.save("inventory_with_total_value.xlsx")






  







